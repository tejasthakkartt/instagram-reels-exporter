import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timezone, timedelta
import threading
import time
import json
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# India Standard Time = UTC+5:30
IST = timezone(timedelta(hours=5, minutes=30))

# Path where session cookies are saved between runs
COOKIES_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ig_session.json")

# ─────────────────────────────────────────────
#  Instagram Internal Web API
# ─────────────────────────────────────────────

IG_APP_ID  = "936619743392459"
IG_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "X-IG-App-ID": IG_APP_ID,
    "X-ASBD-ID":   "129477",
    "Accept":       "*/*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer":      "https://www.instagram.com/",
    "Origin":       "https://www.instagram.com",
}


# ─────────────────────────────────────────────
#  Cookie persistence helpers
# ─────────────────────────────────────────────

def save_cookies(cookies: dict):
    """Persist cookies to disk so the user stays logged in."""
    try:
        with open(COOKIES_FILE, "w") as f:
            json.dump(cookies, f)
    except Exception:
        pass


def load_session_from_file() -> requests.Session | None:
    """
    Try to restore a saved session from disk.
    Returns an authenticated Session if the file exists and has a sessionid,
    otherwise returns None.
    """
    if not os.path.exists(COOKIES_FILE):
        return None
    try:
        with open(COOKIES_FILE) as f:
            cks = json.load(f)
        if not cks.get("sessionid"):
            return None
        s = requests.Session()
        s.headers.update(IG_HEADERS)
        for k, v in cks.items():
            s.cookies.set(k, v, domain=".instagram.com")
        csrf = cks.get("csrftoken", "")
        if csrf:
            s.headers["X-CSRFToken"] = csrf
        return s
    except Exception:
        return None


# ─────────────────────────────────────────────
#  Browser-based login via Selenium
# ─────────────────────────────────────────────

def get_session_via_browser(status_cb=None) -> requests.Session:
    """
    Open a Chrome window, wait for the user to log in to Instagram,
    then extract the session cookies and return an authenticated
    requests.Session.
    """
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.by import By
    from webdriver_manager.chrome import ChromeDriverManager

    if status_cb:
        status_cb("Opening Chrome browser — please log in to Instagram…")

    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=opts,
    )

    try:
        driver.get("https://www.instagram.com/")

        # Wait until user is logged in (sessionid cookie appears)
        if status_cb:
            status_cb("Waiting for Instagram login… (log in in the browser, then this will continue automatically)")

        deadline = time.time() + 180   # 3-minute timeout
        while time.time() < deadline:
            cks = {c["name"]: c["value"] for c in driver.get_cookies()}
            if cks.get("sessionid"):
                break
            time.sleep(1)
        else:
            raise RuntimeError(
                "Login timed out after 3 minutes. Please try again."
            )

        if status_cb:
            status_cb("Login detected — saving session…")

        # Build a requests session with those cookies
        cks = {c["name"]: c["value"] for c in driver.get_cookies()}
        save_cookies(cks)          # ← persist to disk

        s = requests.Session()
        s.headers.update(IG_HEADERS)
        for k, v in cks.items():
            s.cookies.set(k, v, domain=".instagram.com")
        csrf = cks.get("csrftoken", "")
        if csrf:
            s.headers["X-CSRFToken"] = csrf
        return s

    finally:
        driver.quit()


def _get_anonymous_session() -> requests.Session:
    """Fall-back: unauthenticated session (public data only)."""
    s = requests.Session()
    s.headers.update(IG_HEADERS)
    try:
        s.get("https://www.instagram.com/", timeout=15)
    except Exception:
        pass
    csrf = s.cookies.get("csrftoken", "")
    if csrf:
        s.headers["X-CSRFToken"] = csrf
    return s


# ─────────────────────────────────────────────
#  Fetch helpers
# ─────────────────────────────────────────────

def get_user_id(session: requests.Session, username: str) -> tuple:
    url  = "https://www.instagram.com/api/v1/users/web_profile_info/"
    resp = session.get(url, params={"username": username}, timeout=15)
    resp.raise_for_status()
    user = resp.json()["data"]["user"]
    return user["id"], user.get("full_name", username)


def _add_item(media, item, start_buf, end_utc, seen, results, source):
    ts_epoch  = media.get("taken_at") or media.get("device_timestamp", 0)
    post_dt   = datetime.fromtimestamp(ts_epoch, tz=timezone.utc)
    shortcode = media.get("code", "")
    if not shortcode or post_dt < start_buf or post_dt > end_utc:
        return False
    if shortcode in seen:
        return False
    seen.add(shortcode)
    play_count = (
        media.get("play_count") or media.get("view_count")
        or item.get("view_count") or 0
    )
    results.append({
        "reel_link":  f"https://www.instagram.com/reel/{shortcode}/",
        "view_count": int(play_count),
        "timestamp":  post_dt.astimezone(IST).strftime("%Y-%m-%d %H:%M IST"),
        "_ts":        ts_epoch,
    })
    return True


def _fetch_clips(session, user_id, start_buf, end_utc, progress_cb, stop_flag, seen, results):
    """clips/user endpoint — Reels tab."""
    max_id = ""
    page   = 0
    while True:
        if stop_flag and stop_flag():
            break
        page += 1
        if progress_cb:
            progress_cb(f"[Reels tab] fetching page {page}…")
        try:
            resp = session.post(
                "https://www.instagram.com/api/v1/clips/user/",
                data={"target_user_id": user_id, "page_size": "50",
                      "max_id": max_id, "include_feed_video": "true"},
                timeout=20,
            )
            resp.raise_for_status()
            data = resp.json()
        except Exception as exc:
            if progress_cb:
                progress_cb(f"[Reels tab] error: {exc}")
            break

        for item in data.get("items", []):
            media = item.get("media", {})
            _add_item(media, item, start_buf, end_utc, seen, results, "clips")
            if progress_cb and results:
                progress_cb(f"[Reels tab] {len(results)} reels found…")

        paging = data.get("paging_info", {})
        if not paging.get("more_available") or not paging.get("max_id"):
            break
        max_id = paging["max_id"]
        time.sleep(0.4)


def _fetch_feed(session, user_id, start_buf, end_utc, progress_cb, stop_flag, seen, results):
    """feed/user endpoint — profile grid (catches grid-posted videos)."""
    max_id = ""
    page   = 0
    while True:
        if stop_flag and stop_flag():
            break
        page += 1
        if progress_cb:
            progress_cb(f"[Grid] fetching page {page}…")
        try:
            resp = session.get(
                f"https://www.instagram.com/api/v1/feed/user/{user_id}/",
                params={"count": "50", "max_id": max_id},
                timeout=20,
            )
            resp.raise_for_status()
            data = resp.json()
        except Exception as exc:
            if progress_cb:
                progress_cb(f"[Grid] error: {exc}")
            break

        items = data.get("items", [])
        if not items:
            break

        oldest = None
        for item in items:
            ts = item.get("taken_at", 0)
            dt = datetime.fromtimestamp(ts, tz=timezone.utc)
            if oldest is None or dt < oldest:
                oldest = dt
            if item.get("media_type") == 2:
                _add_item(item, item, start_buf, end_utc, seen, results, "grid")
                if progress_cb and results:
                    progress_cb(f"[Grid] {len(results)} reels found…")

        if oldest and oldest < start_buf:
            break
        nxt = data.get("next_max_id", "")
        if not nxt:
            break
        max_id = nxt
        time.sleep(0.4)


def fetch_all_reels(username, start_dt, end_dt,
                    progress_cb=None, stop_flag=None,
                    session: requests.Session = None) -> list:
    """
    Fetch all reels in [start_dt, end_dt] (IST-aware) using both
    the Reels tab and grid feed APIs.
    """
    if start_dt.tzinfo is None:
        start_dt = start_dt.replace(tzinfo=IST)
    if end_dt.tzinfo is None:
        end_dt = end_dt.replace(tzinfo=IST)

    start_utc = start_dt.astimezone(timezone.utc)
    end_utc   = end_dt.astimezone(timezone.utc)
    start_buf = start_utc - timedelta(hours=1)   # tiny buffer for edge cases

    if session is None:
        session = _get_anonymous_session()

    if progress_cb:
        progress_cb("Resolving Instagram User ID…")
    user_id, full_name = get_user_id(session, username)
    if progress_cb:
        progress_cb(f"Found @{full_name} — fetching reels…")

    seen    = set()
    results = []

    _fetch_clips(session, user_id, start_buf, end_utc,
                 progress_cb, stop_flag, seen, results)
    if progress_cb:
        progress_cb(f"Reels tab: {len(results)} reels — scanning grid…")

    _fetch_feed(session, user_id, start_buf, end_utc,
                progress_cb, stop_flag, seen, results)
    if progress_cb:
        progress_cb(f"Done — {len(results)} total reels.")

    results.sort(key=lambda r: r["_ts"], reverse=True)
    for r in results:
        r.pop("_ts", None)
    return results


# ─────────────────────────────────────────────
#  View count K-formatter
# ─────────────────────────────────────────────

def fmt_views(n: int) -> str:
    """Format view count always in k: 900→0.9k, 1420→1.4k, 101000→101k"""
    s = f"{n / 1000:.1f}"
    if s.endswith(".0"):
        s = s[:-2]          # 101.0 → 101
    return s + "k"


# ─────────────────────────────────────────────
#  Excel export
# ─────────────────────────────────────────────

def export_to_excel(rows: list, path: str, include_date: bool = True):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reels"

    thin   = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Set column widths (no header row)
    col_widths = [20, 55] + ([25] if include_date else [])
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    link_font = Font(color="0563C1", underline="single")
    alt_fill  = PatternFill("solid", fgColor="FFF0F5")

    for row_idx, row in enumerate(rows, start=1):   # data starts at row 1
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        # Col 1: View Count (K format)
        c1 = ws.cell(row=row_idx, column=1, value=fmt_views(row["view_count"]))
        c1.alignment = Alignment(horizontal="center")
        c1.border = border; c1.fill = fill
        # Col 2: Reel Link
        c2 = ws.cell(row=row_idx, column=2, value=row["reel_link"])
        c2.hyperlink = row["reel_link"]; c2.font = link_font
        c2.border = border; c2.fill = fill
        # Col 3: Date (optional)
        if include_date:
            c3 = ws.cell(row=row_idx, column=3, value=row["timestamp"])
            c3.alignment = Alignment(horizontal="center")
            c3.border = border; c3.fill = fill

    wb.save(path)


# ─────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Instagram Reels → Excel Exporter")
        self.geometry("740x700")
        self.resizable(False, False)
        self.configure(bg="#0F0F0F")
        self._rows          = []
        self._stop          = False
        self._session       = None
        self.include_date_var = tk.BooleanVar(value=True)
        self._build_ui()
        # Auto-restore saved login session
        self.after(100, self._try_restore_session)

    def _build_ui(self):
        CARD   = "#1A1A1A"
        ACCENT = "#E1306C"
        FG     = "#F5F5F5"
        MUTED  = "#888888"
        ENTRY  = "#242424"
        GREEN  = "#28A745"

        # Banner
        banner = tk.Frame(self, bg=ACCENT, height=72)
        banner.pack(fill="x")
        tk.Label(banner, text="📸  Instagram Reels Exporter",
                 font=("Segoe UI", 18, "bold"),
                 bg=ACCENT, fg="white").pack(pady=20)

        # Card
        card = tk.Frame(self, bg=CARD, padx=28, pady=20)
        card.pack(fill="both", expand=True, padx=20, pady=14)
        card.columnconfigure(0, weight=1)

        def lbl(parent, text, size=10, color=FG, bold=False):
            return tk.Label(parent, text=text, bg=parent["bg"],
                            fg=color, font=("Segoe UI", size, "bold" if bold else "normal"))

        def entry_w(parent, width=None):
            e = tk.Entry(parent, bg=ENTRY, fg=FG, insertbackground=FG,
                         relief="flat", font=("Segoe UI", 12),
                         highlightthickness=1, highlightcolor=ACCENT,
                         highlightbackground="#333")
            if width:
                e.config(width=width)
            return e

        def mk_btn(parent, text, cmd, color, disabled=False):
            return tk.Button(parent, text=text, command=cmd,
                             bg=color, fg="white", relief="flat",
                             font=("Segoe UI", 10, "bold"),
                             padx=14, pady=7, cursor="hand2",
                             activebackground=color, activeforeground="white",
                             state="disabled" if disabled else "normal")

        # ── Step 1: Login ─────────────────────
        lbl(card, "Step 1 — Instagram Login", bold=True, size=11
            ).grid(row=0, column=0, sticky="w", pady=(0, 4))

        sf = tk.Frame(card, bg=CARD)
        sf.grid(row=1, column=0, sticky="ew", pady=(0, 4))

        self.login_btn = mk_btn(sf, "🌐  Open Browser & Login", self._on_login, ACCENT)
        self.login_btn.pack(side="left", padx=(0, 12))

        self.login_status = tk.Label(sf, text="Not logged in",
                                     bg=CARD, fg="#FF6B6B",
                                     font=("Segoe UI", 9, "bold"))
        self.login_status.pack(side="left")

        lbl(card, "Login is saved automatically — you only need to do this once.",
            size=8, color=MUTED).grid(row=2, column=0, sticky="w", pady=(0, 14))

        # Separator
        sep = tk.Frame(card, bg="#333333", height=1)
        sep.grid(row=3, column=0, sticky="ew", pady=(0, 12))

        # ── Step 2: Username ──────────────────
        lbl(card, "Step 2 — Instagram Username  (no @ sign)", bold=True
            ).grid(row=4, column=0, sticky="w", pady=(0, 3))
        self.user_entry = entry_w(card)
        self.user_entry.insert(0, "bannedscenes")
        self.user_entry.grid(row=5, column=0, sticky="ew", ipady=7, pady=(0, 12))

        # ── Step 3: Date range ───────────────
        lbl(card, "Step 3 — Date Range (IST)", bold=True
            ).grid(row=6, column=0, sticky="w", pady=(0, 3))

        df = tk.Frame(card, bg=CARD)
        df.grid(row=7, column=0, sticky="w", pady=(0, 4))

        lbl(df, "From (YYYY-MM-DD)").pack(side="left", padx=(0, 6))
        self.start_entry = entry_w(df, width=14)
        self.start_entry.insert(0, "2026-01-01")
        self.start_entry.pack(side="left", ipady=5)

        lbl(df, "   To").pack(side="left", padx=(12, 6))
        self.end_entry = entry_w(df, width=14)
        self.end_entry.insert(0, datetime.now(IST).strftime("%Y-%m-%d"))
        self.end_entry.pack(side="left", ipady=5)

        lbl(card, "Dates are in IST (India Standard Time). All results will show IST times.",
            size=8, color=MUTED).grid(row=8, column=0, sticky="w", pady=(2, 8))

        # ── Date column checkbox ───────────────
        chk_frame = tk.Frame(card, bg=CARD)
        chk_frame.grid(row=9, column=0, sticky="w", pady=(0, 10))
        tk.Checkbutton(
            chk_frame,
            text=" Include Date Posted column in Excel download",
            variable=self.include_date_var,
            bg=CARD, fg=FG, selectcolor=CARD,
            activebackground=CARD, activeforeground=FG,
            font=("Segoe UI", 10),
            cursor="hand2",
        ).pack(side="left")

        # ── Buttons ───────────────────────────
        bf = tk.Frame(card, bg=CARD)
        bf.grid(row=10, column=0, sticky="ew", pady=(0, 8))

        self.fetch_btn = mk_btn(bf, "🔍  Fetch Reels", self._on_fetch, "#5B6EAE", disabled=True)
        self.fetch_btn.pack(side="left", padx=(0, 10))

        self.stop_btn = mk_btn(bf, "⏹  Stop", self._on_stop, "#555", disabled=True)
        self.stop_btn.pack(side="left", padx=(0, 10))

        self.dl_btn = mk_btn(bf, "📥  Download Excel", self._on_download, GREEN, disabled=True)
        self.dl_btn.pack(side="left")

        # ── Status ────────────────────────────
        self.status_var = tk.StringVar(value="Step 1: Click 'Open Browser & Login' to authenticate.")
        tk.Label(card, textvariable=self.status_var,
                 bg=CARD, fg=MUTED, font=("Segoe UI", 9),
                 wraplength=660, justify="left"
                 ).grid(row=11, column=0, sticky="w")

        # ── Progress ──────────────────────────
        self.progress = ttk.Progressbar(card, mode="indeterminate", length=660)
        self.progress.grid(row=12, column=0, sticky="ew", pady=(5, 8))
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TProgressbar", troughcolor="#242424", background=ACCENT, thickness=6)
        style.configure("Treeview",
                         background="#1A1A1A", fieldbackground="#1A1A1A",
                         foreground=FG, rowheight=26, font=("Segoe UI", 9))
        style.configure("Treeview.Heading",
                         background="#242424", foreground=FG,
                         font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Treeview", background=[("selected", ACCENT)])

        # ── Table ─────────────────────────────
        tf = tk.Frame(card, bg=CARD)
        tf.grid(row=12, column=0, sticky="nsew")
        card.rowconfigure(12, weight=1)

        cols = ("view_count", "reel_link", "date")
        self.tree = ttk.Treeview(tf, columns=cols, show="headings", height=9)
        self.tree.heading("view_count", text="View Count (plays)")
        self.tree.heading("reel_link",  text="Reel Link")
        self.tree.heading("date",       text="Date Posted (IST)")
        self.tree.column("view_count", width=120, anchor="center")
        self.tree.column("reel_link",  width=360, anchor="w")
        self.tree.column("date",       width=150, anchor="center")

        vsb = ttk.Scrollbar(tf, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # ── Summary ───────────────────────────
        self.summary_var = tk.StringVar(value="")
        tk.Label(card, textvariable=self.summary_var,
                 bg=CARD, fg=ACCENT, font=("Segoe UI", 10, "bold")
                 ).grid(row=14, column=0, sticky="w", pady=(6, 0))

    # ── Login ──────────────────────────────────

    def _try_restore_session(self):
        """Called once on startup — silently restore a saved session if available."""
        session = load_session_from_file()
        if session:
            self._session = session
            self.login_status.config(text="✅ Logged in (saved session)", fg="#28A745")
            self.fetch_btn.config(state="normal")
            self._set_status("Session restored! Set your date range and click Fetch Reels.")
        else:
            self._set_status("Step 1: Click 'Open Browser & Login' to authenticate.")

    def _on_login(self):
        self.login_btn.config(state="disabled")
        self.login_status.config(text="Opening browser…", fg="#FFA500")
        self._set_status("Opening Chrome — please log in to Instagram in the browser window…")
        self.progress.start(10)
        threading.Thread(target=self._login_worker, daemon=True).start()

    def _login_worker(self):
        try:
            session = get_session_via_browser(status_cb=self._set_status)
            self._session = session
            self.after(0, self._login_done, True)
        except Exception as e:
            self.after(0, self._login_done, False, str(e))

    def _login_done(self, success: bool, err: str = ""):
        self.progress.stop()
        self.login_btn.config(state="normal")
        if success:
            self.login_status.config(text="✅ Logged in", fg="#28A745")
            self.fetch_btn.config(state="normal")
            self._set_status("Logged in! Set your date range and click Fetch Reels.")
        else:
            self.login_status.config(text="❌ Login failed", fg="#FF6B6B")
            self._set_status(f"Login error: {err}")
            messagebox.showerror("Login Failed", err)

    # ── Fetch ──────────────────────────────────

    def _on_fetch(self):
        username = self.user_entry.get().strip().lstrip("@")
        start_s  = self.start_entry.get().strip()
        end_s    = self.end_entry.get().strip()

        if not username:
            messagebox.showerror("Error", "Please enter an Instagram username.")
            return

        try:
            start_dt = datetime.strptime(start_s, "%Y-%m-%d").replace(tzinfo=IST)
            end_dt   = datetime.strptime(end_s, "%Y-%m-%d").replace(
                hour=23, minute=59, second=59, tzinfo=IST)
        except ValueError:
            messagebox.showerror("Invalid Date", "Dates must be YYYY-MM-DD format.")
            return

        self._rows = []
        self._stop = False
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.summary_var.set("")
        self.fetch_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.dl_btn.config(state="disabled")
        self.progress.start(10)

        threading.Thread(
            target=self._fetch_worker,
            args=(username, start_dt, end_dt),
            daemon=True,
        ).start()

    def _fetch_worker(self, username, start_dt, end_dt):
        try:
            rows = fetch_all_reels(
                username=username,
                start_dt=start_dt,
                end_dt=end_dt,
                progress_cb=self._set_status,
                stop_flag=lambda: self._stop,
                session=self._session,
            )
            self._rows = rows
            self.after(0, self._populate_table, rows)
        except Exception as e:
            self.after(0, self._show_error, str(e))
        finally:
            self.after(0, self._fetch_done)

    def _populate_table(self, rows):
        for r in rows:
            self.tree.insert("", "end", values=(
                fmt_views(r["view_count"]), r["reel_link"], r["timestamp"]))
        total = sum(r["view_count"] for r in rows)
        note  = " (stopped early)" if self._stop else ""
        self.summary_var.set(
            f"✅  {len(rows)} reels{note}  |  Total plays: {fmt_views(total)}"
        )

    def _fetch_done(self):
        self.progress.stop()
        self.fetch_btn.config(state="normal")
        self.stop_btn.config(state="disabled")
        if self._rows:
            self.dl_btn.config(state="normal")
            self._set_status("Done! Click Download Excel to save your file.")
        else:
            self._set_status("No reels found in the selected date range.")

    def _on_stop(self):
        self._stop = True
        self._set_status("Stopping…")
        self.stop_btn.config(state="disabled")

    def _on_download(self):
        if not self._rows:
            messagebox.showinfo("No Data", "Fetch reels first.")
            return
        include_date = self.include_date_var.get()
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile=f"reels_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        if not path:
            return
        try:
            export_to_excel(self._rows, path, include_date=include_date)
            messagebox.showinfo("Saved", f"File saved:\n{path}")
            os.startfile(path)
        except Exception as e:
            messagebox.showerror("Save Error", str(e))

    def _set_status(self, msg: str):
        self.after(0, lambda: self.status_var.set(msg))

    def _show_error(self, msg: str):
        messagebox.showerror("Error", msg)
        self._set_status(f"Error: {msg}")


# ─────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()
