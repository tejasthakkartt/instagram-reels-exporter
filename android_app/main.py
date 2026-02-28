"""
Instagram Reels Exporter — Android App
Built with Kivy for Android
"""
from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen, SlideTransition
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.checkbox import CheckBox
from kivy.uix.progressbar import ProgressBar
from kivy.uix.gridlayout import GridLayout
from kivy.clock import Clock, mainthread
from kivy.utils import platform
from kivy.metrics import dp
from datetime import datetime, timezone, timedelta
import threading
import time
import os
import io

# ── Constants ────────────────────────────────────────────────────────────────
IST      = timezone(timedelta(hours=5, minutes=30))
IG_APP_ID = "936619743392459"
IG_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Linux; Android 13) AppleWebKit/537.36 Chrome/122.0.0.0 Mobile Safari/537.36",
    "X-IG-App-ID": IG_APP_ID,
    "X-ASBD-ID":   "129477",
    "Accept":       "*/*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer":      "https://www.instagram.com/",
}

ACCENT  = (0.882, 0.188, 0.424, 1)   # Instagram pink
DARK    = (0.063, 0.063, 0.063, 1)
CARD    = (0.102, 0.102, 0.102, 1)
FG      = (0.961, 0.961, 0.961, 1)
MUTED   = (0.533, 0.533, 0.533, 1)
GREEN   = (0.157, 0.655, 0.271, 1)

# ── Helpers ──────────────────────────────────────────────────────────────────

def fmt_views(n: int) -> str:
    s = f"{n / 1000:.1f}"
    if s.endswith(".0"):
        s = s[:-2]
    return s + "k"


def _session():
    import requests
    s = requests.Session()
    s.headers.update(IG_HEADERS)
    return s


def _get_session_id():
    """Read saved sessionid from app storage."""
    path = _session_file()
    if os.path.exists(path):
        with open(path) as f:
            return f.read().strip()
    return ""


def _save_session_id(sid: str):
    with open(_session_file(), "w") as f:
        f.write(sid.strip())


def _session_file():
    if platform == "android":
        from android.storage import app_storage_path
        return os.path.join(app_storage_path(), "ig_session.txt")
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "ig_session.txt")


def _downloads_dir():
    if platform == "android":
        return "/storage/emulated/0/Download"
    return os.path.expanduser("~/Downloads")


def get_user_id(session, username):
    resp = session.get(
        "https://www.instagram.com/api/v1/users/web_profile_info/",
        params={"username": username}, timeout=15)
    resp.raise_for_status()
    user = resp.json()["data"]["user"]
    return user["id"], user.get("full_name", username)


def _add_item(media, item, start_buf, end_utc, seen, results):
    ts    = media.get("taken_at") or media.get("device_timestamp", 0)
    dt    = datetime.fromtimestamp(ts, tz=timezone.utc)
    code  = media.get("code", "")
    if not code or dt < start_buf or dt > end_utc or code in seen:
        return
    seen.add(code)
    plays = (media.get("play_count") or media.get("view_count")
             or item.get("view_count") or 0)
    results.append({
        "link":      f"https://www.instagram.com/reel/{code}/",
        "views":     int(plays),
        "timestamp": dt.astimezone(IST).strftime("%Y-%m-%d %H:%M IST"),
        "_ts":       ts,
    })


def fetch_reels(session, username, start_dt, end_dt, progress_cb=None, stop_flag=None):
    if start_dt.tzinfo is None:
        start_dt = start_dt.replace(tzinfo=IST)
    if end_dt.tzinfo is None:
        end_dt = end_dt.replace(tzinfo=IST)
    start_utc = start_dt.astimezone(timezone.utc)
    end_utc   = end_dt.astimezone(timezone.utc)
    start_buf = start_utc - timedelta(hours=1)

    if progress_cb:
        progress_cb("Resolving username…")
    uid, name = get_user_id(session, username)
    if progress_cb:
        progress_cb(f"Found @{name} — fetching…")

    seen, results = set(), []

    # Source 1: clips/user
    max_id, page = "", 0
    while True:
        if stop_flag and stop_flag():
            break
        page += 1
        if progress_cb:
            progress_cb(f"[Reels] page {page} — {len(results)} found")
        try:
            r = session.post(
                "https://www.instagram.com/api/v1/clips/user/",
                data={"target_user_id": uid, "page_size": "50",
                      "max_id": max_id, "include_feed_video": "true"},
                timeout=20)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            if progress_cb:
                progress_cb(f"[Reels] error: {e}")
            break
        for item in data.get("items", []):
            _add_item(item.get("media", {}), item, start_buf, end_utc, seen, results)
        pi = data.get("paging_info", {})
        if not pi.get("more_available") or not pi.get("max_id"):
            break
        max_id = pi["max_id"]
        time.sleep(0.4)

    # Source 2: feed/user
    max_id = ""
    while True:
        if stop_flag and stop_flag():
            break
        try:
            r = session.get(
                f"https://www.instagram.com/api/v1/feed/user/{uid}/",
                params={"count": "50", "max_id": max_id}, timeout=20)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            if progress_cb:
                progress_cb(f"[Grid] error: {e}")
            break
        items = data.get("items", [])
        if not items:
            break
        oldest = None
        for it in items:
            dt = datetime.fromtimestamp(it.get("taken_at", 0), tz=timezone.utc)
            if oldest is None or dt < oldest:
                oldest = dt
            if it.get("media_type") == 2:
                _add_item(it, it, start_buf, end_utc, seen, results)
        if oldest and oldest < start_buf:
            break
        nxt = data.get("next_max_id", "")
        if not nxt:
            break
        max_id = nxt
        time.sleep(0.4)

    results.sort(key=lambda r: r["_ts"], reverse=True)
    for r in results:
        r.pop("_ts", None)
    if progress_cb:
        progress_cb(f"Done — {len(results)} reels found")
    return results


def export_excel(rows, include_date=True) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reels"
    thin   = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    col_widths = [18, 55] + ([25] if include_date else [])
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    link_font = Font(color="0563C1", underline="single")
    alt_fill  = PatternFill("solid", fgColor="FFF0F5")
    for ri, row in enumerate(rows, 1):
        fill = alt_fill if ri % 2 == 0 else PatternFill()
        c1 = ws.cell(ri, 1, fmt_views(row["views"]));   c1.alignment = Alignment(horizontal="center"); c1.border = border; c1.fill = fill
        c2 = ws.cell(ri, 2, row["link"]);   c2.hyperlink = row["link"]; c2.font = link_font; c2.border = border; c2.fill = fill
        if include_date:
            c3 = ws.cell(ri, 3, row["timestamp"]); c3.alignment = Alignment(horizontal="center"); c3.border = border; c3.fill = fill
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── UI Helpers ───────────────────────────────────────────────────────────────

def styled_btn(text, on_press, color=None, disabled=False):
    color = color or ACCENT
    btn = Button(
        text=text, size_hint_y=None, height=dp(50),
        background_color=color, background_normal="",
        font_size=dp(15), bold=True, color=FG,
        disabled=disabled,
    )
    btn.bind(on_press=on_press)
    return btn


def styled_input(hint="", password=False, multiline=False):
    return TextInput(
        hint_text=hint, password=password,
        multiline=multiline, size_hint_y=None, height=dp(48),
        background_color=CARD, foreground_color=FG,
        hint_text_color=MUTED, cursor_color=ACCENT,
        font_size=dp(14), padding=(dp(12), dp(12)),
    )


def section_label(text):
    return Label(
        text=text, size_hint_y=None, height=dp(32),
        color=FG, font_size=dp(13), bold=True,
        halign="left", valign="middle", text_size=(None, dp(32)),
    )


def muted_label(text):
    return Label(
        text=text, size_hint_y=None, height=dp(22),
        color=MUTED, font_size=dp(11),
        halign="left", valign="middle", text_size=(None, dp(22)),
    )


# ── Screens ──────────────────────────────────────────────────────────────────

class LoginScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        root = BoxLayout(orientation="vertical", padding=dp(20), spacing=dp(10),
                         size_hint=(1, 1))
        root.canvas.before.add_in_group = lambda *a: None

        # Header
        hdr = Label(text="[b]📸 Instagram Reels[/b]", markup=True,
                    font_size=dp(22), color=ACCENT,
                    size_hint_y=None, height=dp(60))
        root.add_widget(hdr)

        root.add_widget(section_label("Login with sessionid cookie"))
        root.add_widget(muted_label("Get it from Chrome → instagram.com → F12 → Application → Cookies"))

        self.sid_input = styled_input("Paste your sessionid here", password=True)
        root.add_widget(self.sid_input)

        saved = _get_session_id()
        if saved:
            self.sid_input.text = saved

        root.add_widget(muted_label("Your session is saved — paste only once"))

        self.status_lbl = Label(text="", color=MUTED, font_size=dp(12),
                                size_hint_y=None, height=dp(30))
        root.add_widget(self.status_lbl)

        btn = styled_btn("✅  Save & Continue", self._on_save)
        root.add_widget(btn)

        root.add_widget(Label())   # spacer
        self.add_widget(root)

    def _on_save(self, *a):
        sid = self.sid_input.text.strip()
        if not sid:
            self.status_lbl.text = "Please paste your sessionid"
            self.status_lbl.color = (1, 0.3, 0.3, 1)
            return
        _save_session_id(sid)
        self.manager.transition = SlideTransition(direction="left")
        self.manager.current = "main"


class MainScreen(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self._rows     = []
        self._stop     = False
        self._inc_date = True

        scroll = ScrollView()
        content = BoxLayout(orientation="vertical", padding=dp(16), spacing=dp(10),
                            size_hint_y=None)
        content.bind(minimum_height=content.setter("height"))

        # Banner
        bann = Label(text="[b]📸  Reels Exporter[/b]", markup=True,
                     font_size=dp(20), color=ACCENT,
                     size_hint_y=None, height=dp(52))
        content.add_widget(bann)

        # Username
        content.add_widget(section_label("Instagram Username"))
        self.user_in = styled_input("bannedscenes  (no @ sign)")
        self.user_in.text = "bannedscenes"
        content.add_widget(self.user_in)

        # Date range
        content.add_widget(section_label("Date Range  (IST)"))
        drow = BoxLayout(size_hint_y=None, height=dp(48), spacing=dp(8))
        today = datetime.now(IST).strftime("%Y-%m-%d")
        self.start_in = styled_input("From  YYYY-MM-DD")
        self.start_in.text = "2026-01-01"
        self.end_in = styled_input("To  YYYY-MM-DD")
        self.end_in.text = today
        drow.add_widget(self.start_in)
        drow.add_widget(self.end_in)
        content.add_widget(drow)
        content.add_widget(muted_label("All dates & times shown in IST (India)"))

        # Include date checkbox
        chk_row = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(8))
        self._chk = CheckBox(active=True, size_hint=(None, None),
                             size=(dp(40), dp(40)), color=ACCENT)
        self._chk.bind(active=lambda _, v: setattr(self, "_inc_date", v))
        chk_row.add_widget(self._chk)
        chk_row.add_widget(Label(text="Include Date column in Excel",
                                 color=FG, font_size=dp(13), halign="left",
                                 valign="middle", text_size=(None, dp(40))))
        content.add_widget(chk_row)

        # Buttons
        btn_row = BoxLayout(size_hint_y=None, height=dp(52), spacing=dp(8))
        self.fetch_btn = styled_btn("🔍 Fetch Reels", self._on_fetch)
        self.stop_btn  = styled_btn("⏹ Stop", self._on_stop,
                                    color=(0.3, 0.3, 0.3, 1), disabled=True)
        btn_row.add_widget(self.fetch_btn)
        btn_row.add_widget(self.stop_btn)
        content.add_widget(btn_row)

        self.dl_btn = styled_btn("📥  Download Excel", self._on_download,
                                 color=GREEN, disabled=True)
        content.add_widget(self.dl_btn)

        # Status
        self.status_lbl = Label(text="Enter username & date range, then tap Fetch.",
                                color=MUTED, font_size=dp(12),
                                size_hint_y=None, height=dp(36),
                                halign="center", valign="middle",
                                text_size=(None, dp(36)))
        content.add_widget(self.status_lbl)

        # Summary
        self.summary_lbl = Label(text="", color=ACCENT, font_size=dp(13),
                                 bold=True, size_hint_y=None, height=dp(32),
                                 halign="center", valign="middle",
                                 text_size=(None, dp(32)))
        content.add_widget(self.summary_lbl)

        # Results table header
        hdr_row = GridLayout(cols=3, size_hint_y=None, height=dp(34), spacing=dp(2))
        for t in ["Views", "Reel Link", "Date (IST)"]:
            hdr_row.add_widget(Label(text=f"[b]{t}[/b]", markup=True,
                                     color=ACCENT, font_size=dp(11)))
        content.add_widget(hdr_row)

        # Results grid
        self.result_grid = GridLayout(cols=3, size_hint_y=None, spacing=dp(2))
        self.result_grid.bind(minimum_height=self.result_grid.setter("height"))
        content.add_widget(self.result_grid)

        # Change session link
        chg = Button(text="[ Change Login ]", size_hint_y=None, height=dp(36),
                     background_color=(0, 0, 0, 0), color=MUTED,
                     font_size=dp(11))
        chg.bind(on_press=self._goto_login)
        content.add_widget(chg)

        scroll.add_widget(content)
        self.add_widget(scroll)

    def _goto_login(self, *a):
        self.manager.transition = SlideTransition(direction="right")
        self.manager.current = "login"

    def _set_status(self, msg):
        Clock.schedule_once(lambda dt: setattr(self.status_lbl, "text", msg))

    def _on_fetch(self, *a):
        username = self.user_in.text.strip().lstrip("@")
        start_s  = self.start_in.text.strip()
        end_s    = self.end_in.text.strip()

        if not username:
            self._set_status("Please enter a username")
            return
        try:
            start_dt = datetime.strptime(start_s, "%Y-%m-%d").replace(tzinfo=IST)
            end_dt   = datetime.strptime(end_s, "%Y-%m-%d").replace(
                hour=23, minute=59, second=59, tzinfo=IST)
        except ValueError:
            self._set_status("Invalid date format — use YYYY-MM-DD")
            return

        sid = _get_session_id()
        if not sid:
            self._set_status("No session — go back and save your sessionid")
            return

        self._rows = []
        self._stop = False
        Clock.schedule_once(lambda dt: self.result_grid.clear_widgets())
        Clock.schedule_once(lambda dt: setattr(self.summary_lbl, "text", ""))
        Clock.schedule_once(lambda dt: self.fetch_btn.__setattr__("disabled", True))
        Clock.schedule_once(lambda dt: self.stop_btn.__setattr__("disabled", False))
        Clock.schedule_once(lambda dt: self.dl_btn.__setattr__("disabled", True))

        threading.Thread(
            target=self._worker,
            args=(sid, username, start_dt, end_dt),
            daemon=True,
        ).start()

    def _worker(self, sid, username, start_dt, end_dt):
        import requests
        session = requests.Session()
        session.headers.update(IG_HEADERS)
        session.cookies.set("sessionid", sid, domain=".instagram.com")
        try:
            rows = fetch_reels(
                session, username, start_dt, end_dt,
                progress_cb=self._set_status,
                stop_flag=lambda: self._stop,
            )
            self._rows = rows
            Clock.schedule_once(lambda dt: self._populate(rows))
        except Exception as e:
            self._set_status(f"Error: {e}")
        finally:
            Clock.schedule_once(lambda dt: self._fetch_done())

    @mainthread
    def _populate(self, rows):
        self.result_grid.clear_widgets()
        cols = 3
        row_height = dp(38)
        self.result_grid.row_default_height = row_height
        self.result_grid.row_force_default  = True
        self.result_grid.height = row_height * len(rows)

        for r in rows:
            for val in [fmt_views(r["views"]), r["link"], r["timestamp"]]:
                lbl = Label(text=val, color=FG, font_size=dp(10),
                            halign="center", valign="middle",
                            text_size=(None, row_height))
                self.result_grid.add_widget(lbl)

        total = sum(r["views"] for r in rows)
        self.summary_lbl.text = f"✅  {len(rows)} reels  |  Total: {fmt_views(total)} plays"

    @mainthread
    def _fetch_done(self):
        self.fetch_btn.disabled = False
        self.stop_btn.disabled  = True
        if self._rows:
            self.dl_btn.disabled = False
            self._set_status("Done! Tap Download Excel to save.")
        else:
            self._set_status("No reels found in this date range.")

    def _on_stop(self, *a):
        self._stop = True
        self._set_status("Stopping…")
        self.stop_btn.disabled = True

    def _on_download(self, *a):
        if not self._rows:
            return
        try:
            data = export_excel(self._rows, include_date=self._inc_date)
            dl   = _downloads_dir()
            os.makedirs(dl, exist_ok=True)
            fname = f"reels_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"
            path  = os.path.join(dl, fname)
            with open(path, "wb") as f:
                f.write(data)
            self._set_status(f"✅ Saved to Downloads/{fname}")
        except Exception as e:
            self._set_status(f"Save error: {e}")


# ── App entry ─────────────────────────────────────────────────────────────────

class ReelsApp(App):
    def build(self):
        from kivy.core.window import Window
        Window.clearcolor = DARK

        sm = ScreenManager()
        sm.add_widget(LoginScreen(name="login"))
        sm.add_widget(MainScreen(name="main"))

        if _get_session_id():
            sm.current = "main"

        return sm


if __name__ == "__main__":
    ReelsApp().run()
